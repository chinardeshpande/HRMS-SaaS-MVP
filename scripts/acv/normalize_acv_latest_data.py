#!/usr/bin/env python3
"""Normalize ACV latest implementation spreadsheets into import-ready JSON.

This script intentionally keeps spreadsheet parsing outside the backend runtime.
The backend importers consume the JSON generated here and apply tenant-safe,
idempotent database changes.
"""

from __future__ import annotations

import argparse
import calendar
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SOURCE_DIR = Path(
    "/Users/chinar.deshpande06/Temp/CL-ACV/ACV-India/HRMS-MVP/"
    "ACV Implementation Data/01-source-files/Latest Data"
)

MONTHS = {name.upper(): index for index, name in enumerate(calendar.month_name) if name}
MONTHS.update({name.upper(): index for index, name in enumerate(calendar.month_abbr) if name})

LEAVE_TYPE_MAP = {
    "CL": "casual",
    "PL": "earned",
}

ATTENDANCE_STATUS_MAP = {
    "P": "present",
    "WO": "weekend",
    "W/O": "weekend",
    "WEEKEND": "weekend",
    "PH": "holiday",
    "HOLIDAY": "holiday",
    "H": "holiday",
    "CL": "on_leave",
    "PL": "on_leave",
    "SL": "on_leave",
    "L": "on_leave",
    "LEAVE": "on_leave",
    "LOP": "absent",
    "A": "absent",
    "ABSENT": "absent",
    "HD": "half_day",
    "HALF": "half_day",
}


def norm(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def parse_float(value: Any) -> float:
    if value is None or norm(value) == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def month_number(value: Any) -> int | None:
    key = norm(value).upper()
    return MONTHS.get(key)


def normalize_leave(source_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    source = source_dir / "Leave_Management_2026.xlsx"
    balances: list[dict[str, Any]] = []
    lop_evidence: list[dict[str, Any]] = []
    warnings: list[str] = []

    if not source.exists():
        warnings.append(f"Leave workbook missing: {source}")
        return balances, lop_evidence, warnings

    workbook = load_workbook(source, data_only=True, read_only=True)
    for worksheet in workbook.worksheets:
        employee_name = worksheet.title
        monthly_rows: list[dict[str, Any]] = []

        for row in worksheet.iter_rows(min_row=4, values_only=True):
            month = month_number(row[0] if len(row) > 0 else None)
            if not month:
                continue

            row_data = {
                "month": month,
                "monthName": calendar.month_name[month],
                "casual": {
                    "total": parse_float(row[1] if len(row) > 1 else None),
                    "taken": parse_float(row[2] if len(row) > 2 else None),
                    "balance": parse_float(row[3] if len(row) > 3 else None),
                },
                "earned": {
                    "total": parse_float(row[4] if len(row) > 4 else None),
                    "taken": parse_float(row[5] if len(row) > 5 else None),
                    "balance": parse_float(row[6] if len(row) > 6 else None),
                },
                "employeeStatus": norm(row[7] if len(row) > 7 else None),
                "lop": parse_float(row[8] if len(row) > 8 else None),
                "remarks": norm(row[9] if len(row) > 9 else None),
            }
            monthly_rows.append(row_data)

            if row_data["lop"]:
                lop_evidence.append(
                    {
                        "sourceFile": str(source),
                        "sourceSheet": worksheet.title,
                        "employeeName": employee_name,
                        "month": month,
                        "monthName": row_data["monthName"],
                        "lopDays": row_data["lop"],
                        "remarks": row_data["remarks"],
                    }
                )

        for leave_code, leave_type in LEAVE_TYPE_MAP.items():
            rows_with_data = [
                row for row in monthly_rows if row[leave_type]["balance"] or row[leave_type]["taken"] or row[leave_type]["total"]
            ]
            if not rows_with_data:
                continue

            final_row = rows_with_data[-1]
            total_used = round(sum(row[leave_type]["taken"] for row in rows_with_data), 2)
            final_balance = round(final_row[leave_type]["balance"], 2)
            total_allocated = round(final_balance + total_used, 2)

            balances.append(
                {
                    "sourceFile": str(source),
                    "sourceSheet": worksheet.title,
                    "employeeName": employee_name,
                    "year": 2026,
                    "leaveCode": leave_code,
                    "leaveType": leave_type,
                    "totalAllocated": total_allocated,
                    "used": total_used,
                    "pending": 0,
                    "carriedForward": 0,
                    "encashed": 0,
                    "available": final_balance,
                    "finalMonth": final_row["month"],
                    "monthlyEvidence": rows_with_data,
                    "normalizationNote": "Imported allocation is computed as final balance plus cumulative monthly taken, preserving available balance from source.",
                }
            )

    return balances, lop_evidence, warnings


def normalize_attendance(source_dir: Path) -> tuple[list[dict[str, Any]], list[str]]:
    attendance_dir = source_dir / "Monthly attendance"
    records: list[dict[str, Any]] = []
    warnings: list[str] = []

    if not attendance_dir.exists():
        warnings.append(f"Monthly attendance directory missing: {attendance_dir}")
        return records, warnings

    for source in sorted(attendance_dir.glob("*.xlsx")):
        workbook = load_workbook(source, data_only=True, read_only=True)
        for worksheet in workbook.worksheets:
            year = None
            if str(worksheet.title).strip().isdigit():
                year = int(str(worksheet.title).strip())

            header = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if not header:
                warnings.append(f"{source.name}/{worksheet.title}: missing day header row")
                continue

            day_columns = []
            for index, value in enumerate(header):
                if isinstance(value, int) and 1 <= value <= 31:
                    day_columns.append((index, value))

            current_employee_name = ""
            current_employee_code = ""
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                row_year = year
                if row_year is None:
                    # Files with a generic sheet name are current 2026 ACV monthly sheets.
                    row_year = 2026

                raw_employee_name = norm(row[1] if len(row) > 1 else None)
                raw_employee_code = norm(row[3] if len(row) > 3 else None)
                employee_name = raw_employee_name or current_employee_name
                employee_code = raw_employee_code if raw_employee_name else current_employee_code
                month = month_number(row[2] if len(row) > 2 else None)

                if raw_employee_name:
                    current_employee_name = employee_name
                if raw_employee_name and raw_employee_code:
                    current_employee_code = employee_code

                if not employee_name or not month:
                    continue

                for index, day in day_columns:
                    if day > calendar.monthrange(row_year, month)[1]:
                        continue
                    raw = norm(row[index] if index < len(row) else None).upper()
                    if not raw:
                        continue
                    status = ATTENDANCE_STATUS_MAP.get(raw)
                    if not status:
                        warnings.append(f"{source.name}/{worksheet.title}: unknown status '{raw}' for {employee_name} {row_year}-{month:02d}-{day:02d}")
                        continue

                    records.append(
                        {
                            "sourceFile": str(source),
                            "sourceSheet": worksheet.title,
                            "employeeName": employee_name,
                            "employeeCode": employee_code,
                            "date": f"{row_year}-{month:02d}-{day:02d}",
                            "year": row_year,
                            "month": month,
                            "day": day,
                            "sourceCode": raw,
                            "status": status,
                            "notes": f"Imported from ACV monthly attendance source ({raw}).",
                        }
                    )

    return records, warnings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR))
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    leave_balances, lop_evidence, leave_warnings = normalize_leave(source_dir)
    attendance_records, attendance_warnings = normalize_attendance(source_dir)

    payload = {
        "generatedAt": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "sourceDir": str(source_dir),
        "summary": {
            "leaveBalanceRows": len(leave_balances),
            "lopEvidenceRows": len(lop_evidence),
            "attendanceRows": len(attendance_records),
            "warnings": len(leave_warnings) + len(attendance_warnings),
        },
        "leaveBalances": leave_balances,
        "lopEvidence": lop_evidence,
        "attendanceRecords": attendance_records,
        "warnings": leave_warnings + attendance_warnings,
    }

    (output_dir / "acv-latest-normalized.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    (output_dir / "README.md").write_text(
        "\n".join(
            [
                "# ACV Latest Data Normalization",
                "",
                f"Generated: {payload['generatedAt']}",
                f"Source directory: `{source_dir}`",
                "",
                "## Summary",
                "",
                f"- Leave balance rows: {len(leave_balances)}",
                f"- LOP evidence rows: {len(lop_evidence)}",
                f"- Attendance rows: {len(attendance_records)}",
                f"- Warnings: {len(payload['warnings'])}",
                "",
                "## Notes",
                "",
                "- Leave balances are normalized from ACV's leave workbook and are not payroll calculations.",
                "- LOP rows are preserved as evidence because ACV currently has no active unpaid leave policy.",
                "- Attendance rows are normalized from monthly attendance sheets only; biometric `.xls` files remain out of scope until a safe legacy parser/converter is selected.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], indent=2))


if __name__ == "__main__":
    main()
